import re
import streamlit as st
import io
import csv
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CIRCLE = ["①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩"]
FONT   = "メイリオ"
FS     = 10

COLORS = {
    "title_bg": "1F3864", "title_fg": "FFFFFF",
    "note_bg":  "D9E1F2",
    "hdr_bg":   "1F3864", "hdr_fg":   "FFFFFF",
    "hdr2_bg":  "2E75B6",
    "v": ["FFF2CC", "E2EFDA", "DDEBF7", "FCE4D6", "EDE7F6", "F3F3F3"],
}

# ── 単位判定 ──────────────────────────────────────────────────────────
def detect_unit(name: str) -> str:
    n = name.strip()
    if "回数" in n:
        return "回数"
    if "日数" in n or n.endswith("D"):
        return "日数"
    if any(k in n for k in ["特所定D","特時外D","特休日D","出勤日","欠勤日","傷病日","事故日","無断日"]):
        return "日数"
    if any(k in n for k in ["エラー","チェック","判定","ミス","乖離","比較","重複","非表示","在宅予定","在宅実績"]):
        return "その他"
    return "時間"

# ── 日本語説明文（ネスト条件解析 + 算術チェーン集約） ───────────────
def generate_description(aid_name: str, step_list: list) -> str:
    if not step_list:
        return f"{aid_name}を算出する（ステップなし）"

    def parse_levels(cond: str) -> dict:
        result = {}
        for m in re.finditer(r'第(\d+)条件レベル[：:]\s*(真|偽)', cond):
            result[int(m.group(1))] = (m.group(2) == '真')
        return result

    COND_KW = frozenset(["CDの検査", "２値の比較", "比較", "判定"])
    has_nested = any('条件レベル' in s.get('cond', '') for s in step_list)

    cond_labels = {}
    calc_ops    = []

    for s in step_list:
        cond   = s.get("cond",  "").strip()
        ptype  = s.get("ptype", "").strip()
        desc   = s.get("desc",  "").strip()
        levels = parse_levels(cond)

        if has_nested and any(kw in ptype for kw in COND_KW):
            next_lvl = max(levels.keys(), default=0) + 1
            cond_labels[next_lvl] = desc
        else:
            calc_ops.append({"levels": levels, "ptype": ptype, "desc": desc})

    if not calc_ops:
        return "\n".join(
            f"({i+1}) [{s.get('cond','')}] {s.get('ptype','')}：{s.get('desc','')}"
            for i, s in enumerate(step_list)
        )

    max_depth = max(len(op["levels"]) for op in calc_ops)
    active    = [op for op in calc_ops if len(op["levels"]) == max_depth]

    if all("固定値0" in op["desc"] for op in active):
        return f"{aid_name}は本パターンでは対象外（常に0）"

    def clean_cond(label: str, truth: bool) -> str:
        label = re.sub(r'（[^）]*）|\([^)]*\)', '', label).strip()
        label = re.sub(r'(の場合真|の場合偽|なら真|なら偽)$', '', label).strip()
        return label if truth else f"{label}ではない"

    cond_parts = []
    if has_nested and active:
        for lvl in sorted(active[0]["levels"].keys()):
            raw   = cond_labels.get(lvl, f"条件{lvl}")
            truth = active[0]["levels"][lvl]
            cond_parts.append(clean_cond(raw, truth))
    elif not has_nested:
        NO_COND_MARKER = re.compile(r'^\[?-+\]?$')
        for s in step_list:
            c = s.get("cond", "").strip()
            if c and not NO_COND_MARKER.match(c):
                cleaned = re.sub(r'^\[.*?\]\s*', '', c).strip()
                cleaned = re.sub(r'（[^）]*）|\([^)]*\)', '', cleaned).strip()
                cleaned = re.sub(r'(の場合真|の場合偽|なら真|なら偽)$', '', cleaned).strip()
                if cleaned:
                    cond_parts.append(cleaned)
                break

    base_val = None
    sub_list = []
    add_list = []

    for op in active:
        raw   = op["desc"].strip()
        ptype = op.get("ptype", "").strip()
        if not raw or "固定値0" in raw:
            continue
        if "固定値1" in raw:
            add_list.append("1（発生フラグ）")
            continue

        cleaned = re.sub(r'^一時変数\d+\s*', '', raw).strip()
        if not cleaned or '一時変数' in cleaned:
            continue

        if ptype in ('加算', '合算'):
            if base_val is None:
                base_val = cleaned
            else:
                add_list.append(cleaned)
            continue
        if ptype == '減算':
            sub_list.append(cleaned)
            continue

        if cleaned and cleaned[0] in ('-', '－', '+', '＋'):
            sign = cleaned[0]
            val  = cleaned[1:].strip()
            if sign in ('-', '－'):
                sub_list.append(val)
            else:
                add_list.append(val)
        elif '－' in cleaned:
            lhs, _, rhs = cleaned.partition('－')
            lhs, rhs = lhs.strip(), rhs.strip()
            if base_val is None and lhs and '一時変数' not in lhs:
                base_val = lhs
            if rhs:
                sub_list.append(rhs)
        elif '＋' in cleaned:
            lhs, _, rhs = cleaned.partition('＋')
            lhs, rhs = lhs.strip(), rhs.strip()
            if base_val is None and lhs and '一時変数' not in lhs:
                base_val = lhs
            if rhs:
                add_list.append(rhs)
        elif '-' in cleaned:
            lhs, _, rhs = cleaned.partition('-')
            lhs, rhs = lhs.strip(), rhs.strip()
            if base_val is None and lhs and '一時変数' not in lhs:
                base_val = lhs
            if rhs:
                sub_list.append(rhs)
        elif '+' in cleaned:
            lhs, _, rhs = cleaned.partition('+')
            lhs, rhs = lhs.strip(), rhs.strip()
            if base_val is None and lhs and '一時変数' not in lhs:
                base_val = lhs
            if rhs:
                add_list.append(rhs)
        elif cleaned and '一時変数' not in cleaned and base_val is None:
            base_val = cleaned

    cond_str = "、".join(cond_parts) + "において" if cond_parts else ""

    if base_val:
        calc = base_val
        if sub_list and add_list:
            calc += f"から{'・'.join(sub_list)}を差し引き、{'・'.join(add_list)}を加算した時間"
        elif sub_list:
            calc += f"から{'・'.join(sub_list)}を差し引いた時間"
        elif add_list:
            calc += f"に{'・'.join(add_list)}を加算した時間"
    else:
        items = sub_list + add_list
        calc  = ("・".join(items) + "の集計値") if items else aid_name

    if cond_str:
        return f"{cond_str}、{calc}を算出する。"
    else:
        return f"{calc}を算出する。"

# ── CSV 読み込み ──────────────────────────────────────────────────────
def decode_csv(raw: bytes):
    for enc in ["utf-8-sig", "utf-8", "cp932", "shift_jis"]:
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return None, None

# ── データ処理 ────────────────────────────────────────────────────────
def process(content: str):
    reader = csv.DictReader(io.StringIO(content))
    rows   = list(reader)
    if not rows:
        return None

    sc_map  = OrderedDict()
    aid_map = OrderedDict()
    steps   = defaultdict(list)

    for r in rows:
        sc  = r.get("集約設定コード",     "").strip()
        scn = r.get("集約設定コード名称", "").strip()
        aid = r.get("集約値ID",           "").strip()
        adn = r.get("集約値ID名称",       "").strip()
        if sc  and sc  not in sc_map:  sc_map[sc]  = scn
        if aid and aid not in aid_map: aid_map[aid] = adn
        if sc and aid:
            steps[(sc, aid)].append({
                "ord":   r.get("実行順序",  "").strip(),
                "cond":  r.get("実行条件",  "").strip(),
                "ptype": r.get("処理タイプ","").strip(),
                "desc":  r.get("処理説明",  "").strip(),
            })

    def skey(x):
        try:    return int(x)
        except: return 9999
    aid_list = sorted(aid_map.keys(), key=skey)
    sc_list  = list(sc_map.keys())

    def normalize(desc: str) -> str:
        d = re.sub(r'一時変数\d+', '一時変数', desc)
        d = re.sub(r'\s+', ' ', d)
        return d.strip()

    def fkey(sc, aid):
        sorted_s = sorted(steps.get((sc, aid), []),
                          key=lambda x: int(x["ord"]) if x["ord"].isdigit() else 999)
        return tuple((s["cond"], s["ptype"], normalize(s["desc"])) for s in sorted_s)

    variants = {}
    for aid in aid_list:
        used_sc = [sc for sc in sc_list if (sc, aid) in steps]
        vmap, sc_v = {}, {}
        for sc in used_sc:
            fk = fkey(sc, aid)
            if fk not in vmap:
                vmap[fk] = CIRCLE[len(vmap)] if len(vmap) < len(CIRCLE) else f"({len(vmap)+1})"
            sc_v[sc] = vmap[fk]
        vdetail = {}
        for fk, vn in vmap.items():
            vdetail[vn] = [sc for sc in used_sc if fkey(sc, aid) == fk]
        variants[aid] = (sc_v, vdetail)

    return dict(sc_map=sc_map, sc_list=sc_list,
                aid_map=aid_map, aid_list=aid_list,
                steps=steps, variants=variants,
                n_rows=len(rows))

# ── スタイルヘルパー ──────────────────────────────────────────────────
def _s(): return Side(style="thin", color="000000")
def _bdr(): s = _s(); return Border(left=s, right=s, top=s, bottom=s)
def _fl(c): return PatternFill("solid", fgColor=c)

def put(ws, row, col, val="", bg=None, fg="000000",
        bold=False, halign="left", wrap=True, sz=None):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = _fl(bg)
    c.font      = Font(name=FONT, color=fg, bold=bold, size=sz or FS)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    c.border    = _bdr()
    return c

# ── Excel 生成 ────────────────────────────────────────────────────────
def build_excel(data: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "パターン一覧"
    ws.sheet_view.showGridLines = False

    sc_map   = data["sc_map"]
    sc_list  = data["sc_list"]
    aid_map  = data["aid_map"]
    aid_list = data["aid_list"]
    steps_d  = data["steps"]
    variants = data["variants"]

    # 固定列: No / ID / 項目名称 / 単位 / パターン / 概要 / 備考  (A〜G)
    N_FIXED = 7
    glabels = [sc_map[sc] for sc in sc_list]
    N_TOTAL = N_FIXED + len(glabels)

    # ── 列幅 ──────────────────────────────────────────────────────────
    col_widths = [5, 10, 20, 6, 6, 70, 14]   # No, ID, 項目名称, 単位, パターン, 概要, 備考
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for gi in range(N_FIXED + 1, N_TOTAL + 1):
        ws.column_dimensions[get_column_letter(gi)].width = 13

    # ── タイトル行 ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 26
    c = ws.cell(1, 1, "算出パターンデータ一覧")
    c.fill      = _fl(COLORS["title_bg"])
    c.font      = Font(name=FONT, color=COLORS["title_fg"], bold=True, size=14)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_TOTAL)

    # ── 注記 ──────────────────────────────────────────────────────────
    notes = [
        "■ 単位：時間 / 日数 / 回数 / その他",
        "■ バリアント記号（①②③…）が同じ行は同一の計算ロジック。パターン列に記号がある場合はそのコードに適用あり。",
        "■ 同じIDで複数行ある場合はパターンごとに計算内容が異なる。",
    ]
    for i, note in enumerate(notes, start=2):
        ws.row_dimensions[i].height = 14
        c = ws.cell(i, 1, note)
        c.fill      = _fl(COLORS["note_bg"])
        c.font      = Font(name=FONT, size=8, italic=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=N_TOTAL)

    # ── ヘッダー1行目 ─────────────────────────────────────────────────
    H1 = 6
    ws.row_dimensions[H1].height = 22
    for ci, h in enumerate(["No", "ID", "項目名称", "単位", "パターン", "概要", "備考"], 1):
        put(ws, H1, ci, h, bg=COLORS["hdr_bg"], fg=COLORS["hdr_fg"],
            bold=True, halign="center")
    for gi, gl in enumerate(glabels, start=N_FIXED + 1):
        put(ws, H1, gi, gl, bg=COLORS["hdr_bg"], fg=COLORS["hdr_fg"],
            bold=True, halign="center")

    ws.freeze_panes = ws.cell(H1 + 1, 1)

    # ── データ行 ──────────────────────────────────────────────────────
    cur = H1 + 1
    no  = 1

    for aid in aid_list:
        aid_name   = aid_map[aid]
        sc_v, vdet = variants[aid]
        unit       = detect_unit(aid_name)

        all_variants = sorted(set(sc_v.values()),
                              key=lambda v: CIRCLE.index(v) if v in CIRCLE else 99)
        if not all_variants:
            continue

        n_rows    = len(all_variants)
        row_start = cur
        row_end   = cur + n_rows - 1

        # A〜D: No / ID / 項目名称 / 単位（複数バリアントは縦結合）
        put(ws, row_start, 1, no,       halign="center")
        put(ws, row_start, 2, int(aid) if aid.isdigit() else aid, halign="center")
        put(ws, row_start, 3, aid_name, halign="left")
        put(ws, row_start, 4, unit,     halign="center")

        if n_rows > 1:
            for col in range(1, 5):
                for r in range(row_start + 1, row_end + 1):
                    put(ws, r, col, "",
                        halign=ws.cell(row_start, col).alignment.horizontal)
                ws.merge_cells(
                    start_row=row_start, start_column=col,
                    end_row=row_end,     end_column=col,
                )

        # E〜: バリアントごとに1行
        for vi_idx, vn in enumerate(all_variants):
            r         = row_start + vi_idx
            rep_sc    = vdet[vn][0]
            step_list = sorted(steps_d.get((rep_sc, aid), []),
                               key=lambda s: int(s["ord"]) if s["ord"].isdigit() else 999)
            summary   = generate_description(aid_name, step_list)

            ws.row_dimensions[r].height = FS + 5

            put(ws, r, 5, vn, halign="center", sz=13)  # E: パターン
            put(ws, r, 6, summary)                                  # F: 概要
            put(ws, r, 7, "")                                       # G: 備考

            for gi, sc in enumerate(sc_list, start=N_FIXED + 1):
                if sc_v.get(sc) == vn:
                    put(ws, r, gi, vn, halign="center")
                else:
                    put(ws, r, gi, "-", halign="center")

        no  += 1
        cur  = row_end + 1

    ws.auto_filter.ref = f"A{H1}:{get_column_letter(N_TOTAL)}{cur - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


APP_NAME = "算出パターンデータ一覧"

# ── Streamlit UI ──────────────────────────────────────────────────────
st.set_page_config(
    page_title=APP_NAME,
    page_icon="📊",
    layout="wide",
)
st.title(f"📊 {APP_NAME}")
st.caption("CSVをアップロード → 集約設定コードごとのパターンマトリクス付きExcelを生成します。")
st.divider()

uploaded = st.file_uploader("CSVをドラッグ or クリックして選択", type=["csv"])

if uploaded:
    raw = uploaded.read()
    content, enc = decode_csv(raw)
    if content is None:
        st.error("文字コードを判定できませんでした。UTF-8 または CP932 のCSVをご確認ください。")
        st.stop()

    with st.spinner("解析中…"):
        data = process(content)
    if data is None:
        st.error("CSVを読み取れませんでした。ヘッダー行を含む正しい形式かご確認ください。")
        st.stop()

    # ── サイドバー: 集約設定コード一覧
    with st.sidebar:
        st.header("📋 集約設定コード一覧")
        st.caption(f"合計 {len(data['sc_list'])} コード")
        st.divider()
        for sc in data["sc_list"]:
            st.markdown(f"**`{sc}`** {data['sc_map'][sc]}")

    # ── メトリクス
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("文字コード",     enc)
    c2.metric("集約設定コード", f"{len(data['sc_list'])} 件")
    c3.metric("集約値ID",       f"{len(data['aid_list'])} 件")
    c4.metric("データ行数",     f"{data['n_rows']} 行")
    st.divider()

    # ── データプレビュー
    st.subheader("📋 データプレビュー")

    preview_rows = []
    no = 1
    for aid in data["aid_list"]:
        aname      = data["aid_map"][aid]
        sc_v, vdet = data["variants"][aid]
        all_v      = sorted(set(sc_v.values()),
                            key=lambda v: CIRCLE.index(v) if v in CIRCLE else 99)
        unit = detect_unit(aname)
        for vn in all_v:
            rep_sc    = vdet[vn][0]
            step_list = sorted(data["steps"].get((rep_sc, aid), []),
                               key=lambda s: int(s["ord"]) if s["ord"].isdigit() else 999)
            desc = generate_description(aname, step_list)
            apply_codes = [data["sc_map"][sc]
                           for sc in data["sc_list"] if sc_v.get(sc) == vn]
            preview_rows.append({
                "No":         no,
                "ID":         aid,
                "項目名称":   aname,
                "単位":       unit,
                "パターン":   vn,
                "計算式概要": desc,
                "適用コード": " / ".join(apply_codes),
            })
        no += 1

    st.dataframe(
        preview_rows,
        use_container_width=True,
        hide_index=True,
        height=420,
        column_config={
            "No":         st.column_config.NumberColumn(width="small"),
            "ID":         st.column_config.TextColumn(width="small"),
            "項目名称":   st.column_config.TextColumn(width="medium"),
            "単位":       st.column_config.TextColumn(width="small"),
            "パターン":   st.column_config.TextColumn(width="small"),
            "計算式概要": st.column_config.TextColumn(width="large"),
            "適用コード": st.column_config.TextColumn(width="large"),
        },
    )
    st.divider()

    # ── Excel 生成 & ダウンロード
    with st.spinner("Excel生成中…"):
        buf     = build_excel(data)
        outname = uploaded.name.replace(".csv", f"_{APP_NAME}.xlsx")

    st.success("✅ 生成完了！下のボタンからダウンロードしてください。")
    st.download_button(
        label="📥 Excelをダウンロード",
        data=buf.getvalue(),
        file_name=outname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

else:
    st.info("👆 CSVファイルをドラッグするか、クリックして選択してください。")
    with st.sidebar:
        st.header("📋 使い方")
        st.markdown("""
1. リアルタイム集約式設定CSVをアップロード
2. 画面のテーブルでデータを確認
3. **Excelをダウンロード**ボタンをクリック

**Excelの出力内容**
- 集約値ID ごとに1行（複数パターンは行分割）
- A〜D列は同一IDで縦結合
- E列：パターン（①②③）
- F列：計算式の日本語概要
- H列〜：集約設定コードごとの適用パターン
""")
